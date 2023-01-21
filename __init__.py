import os
import olefile
import zlib
import struct
from collections import OrderedDict
import openpyxl
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Color, fills, Side
from copy import copy

def get_hwp_text(filename):
    f = olefile.OleFileIO(filename)
    dirs = f.listdir()

    # HWP 파일 검증
    if ["FileHeader"] not in dirs or \
       ["\x05HwpSummaryInformation"] not in dirs:
        raise Exception("Not Valid HWP.")

    # 문서 포맷 압축 여부 확인
    header = f.openstream("FileHeader")
    header_data = header.read()
    is_compressed = (header_data[36] & 1) == 1

    # Body Sections 불러오기
    nums = []
    for d in dirs:
        if d[0] == "BodyText":
            nums.append(int(d[1][len("Section"):]))
    sections = ["BodyText/Section"+str(x) for x in sorted(nums)]

    # 예외 처리 
    bad_bytes = [
        '\x0b漠杳\x00\x00\x00\x00\x0b',
        '\x0b氠瑢\x00\x00\x00\x00\x0b',
        '\x15湯湷\x00\x00\x00\x00\x15',
        '\U000f0288'
    ]

    # 전체 text 추출
    text = ""
    for section in sections:
        bodytext = f.openstream(section)
        data = bodytext.read()
        if is_compressed:
            unpacked_data = zlib.decompress(data, -15)
        else:
            unpacked_data = data
    
        # 각 Section 내 text 추출    
        section_text = ""
        i = 0
        size = len(unpacked_data)
        while i < size:
            header = struct.unpack_from("<I", unpacked_data, i)[0]
            rec_type = header & 0x3ff
            rec_len = (header >> 20) & 0xfff

            if rec_type in [67]:
                rec_data = unpacked_data[i+4:i+4+rec_len]
                decode_rec = rec_data.decode('utf-16')
                for bad in bad_bytes :
                    if bad in decode_rec :
                        decode_rec = ''

                if not decode_rec == '' : 
                    section_text += decode_rec
                    section_text += "\n"
                else :
                    section_text += '-'*16
                    section_text += "\r\n"

            i += 4 + rec_len

        text += section_text
        text += "\n"

    f.close()
    
    return text

def pcsi_setting(survey_name='', 
                division='', 
                info_text_key = '면접원 지시사항', 
                da='정윤교',
                qnr_folder = 'QNR', 
                save_folder = 'SET',) :

    if survey_name == '' or not type(survey_name) == str :
        print('❌ ERROR : 기관명은 문자형으로 입력')
        return

    if division == '' or not type(division) == str or not division in ['KMAC', 'KSA']:
        print('❌ ERROR : 구분은 KMAC/KSA로만 입력 (대소문자 정확하게)')
        return

    curr_files = os.listdir(qnr_folder)
    curr_files = {int(i.split('.')[0]):i for i in curr_files if '.hwp' in i}
    sort_files = sorted(curr_files.items())
    
    hwps = [hwp for key, hwp in sort_files]

    key_texts = []
    if division == 'KMAC' :
        key_texts = ['SQ3', 'SQ4']

    if division == 'KSA' : 
        key_texts = ['SQ3', 'SQ4', '문 9', '문 10', '문 12']
    # 워딩이 다른 문항
    type_code = {
        'A' : 1,
        'B' : 2,
        'C' : 3,
        'D' : 4,
        'E' : 5,
        'F' : 6,
        'G' : 7,
        'H' : 8
    }

    da_num = {
        '정윤교' : 7353,
        '문영선' : 7311,
        '고혜경' : 7267,
        '조성희' : 7314,
        '전미림' : 7228,
        '정미지' : 7189,
    }


    change_cells = OrderedDict()
    for key in key_texts :
        change_cells[key] = OrderedDict()

    change_cells['qnrs'] = OrderedDict()
    change_cells['info'] = OrderedDict()
    change_cells['Q8'] = OrderedDict()


    for hwp in hwps :
        # QNR 세팅
        del_hwp = hwp.replace('.hwp', '')
        code, label = del_hwp.split('.')
        name, qtype = label.split('_')
        change_cells['qnrs'][code] = {'name': name, 'type': qtype, 'type_code': type_code[qtype]}
        
        # SQ 세팅 (공통)
        curr_hwp = get_hwp_text(os.path.join(os.getcwd(), qnr_folder, hwp)).split('\r\n')
        for key in key_texts :
            curr_txt = [i for i in curr_hwp if key in i]
            if not curr_txt :
                continue
            set_word = curr_txt[0]
            set_word = set_word.replace('○○', '고객')
            set_word = set_word.replace(f'{key}. ', '')
            set_word = set_word.replace(f'{key}】', '')
            set_word = set_word.replace(f'･', '·')
            if '】' in set_word :
              set_word = set_word.split('】')
              set_word = ''.join(set_word[1:])
            
            set_word = set_word.replace(survey_name, f'<font color=blue>{survey_name}</font>')
            change_cells[key][code] = set_word.strip()
        
        # KSA ONLY
        if division == 'KSA' :
            Q8_array = []
            for idx, tx in enumerate(curr_hwp) :
                if '8-' in tx and not '】' in tx:
                  chk_next = curr_hwp[idx+1]
                  if not ('실사 책임자' in chk_next or '0000' in chk_next) :
                    Q8_array.append(chk_next)

            change_cells['Q8'][code] = Q8_array

        # SQ 이 후 조사 시작전 안내 문구
        info_txt = []
        info_flag = False
        for tx in curr_hwp :
            if not info_text_key and not info_text_key in tx :
                continue
            if info_text_key in tx :
                info_flag = True
                continue
            
            if info_flag and '-'*16 == tx :
                break

            if info_flag :
                set_word = tx
                set_word = set_word.replace(survey_name, f'<font color=blue>{survey_name}</font>')
                info_txt.append(set_word.strip())
        
        info_txt = '<br/><br/>'.join(info_txt)
        info_txt = info_txt.replace('<br/><br/><br/><br/>', '<br/><br/>')
        info_txt = info_txt.replace(f'･', '·')
        change_cells['info'][code] = info_txt
        

    wb = openpyxl.load_workbook('template.xlsx')
    wb_sheetname = wb.sheetnames[0]
    ws = wb[wb_sheetname]
    rows = ws.rows
    cols = ws.columns

    new_wb = openpyxl.Workbook()
    new_wb.active.title = wb_sheetname
    new_ws = new_wb.active

    for row in rows :
        for cell in row :
            curr_cell = new_ws.cell(row=cell.row, column=cell.column)
            curr_cell.value = cell.value
            if cell.has_style :
                curr_cell.font = copy(cell.font)
                curr_cell.border = copy(cell.border)
                curr_cell.fill = copy(cell.fill)
                curr_cell.number_format = copy(cell.number_format)
                curr_cell.protection = copy(cell.protection)
                curr_cell.alignment = copy(cell.alignment)

    use_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    for us in use_columns :
        new_ws.column_dimensions[us].width = ws.column_dimensions[us].width


    # 기관명 세팅 관련
    name_set = new_ws.cell(27, 6)
    name_set.value = name_set.value%(survey_name)

    last_text = {
        'KMAC' : '''<div align=center style='border: 1px solid #800000;padding:10px; white-space: normal; font-size:13px; border-collapse: collapse;background-color:#5579d7;color:#ffffff'>
▣ 지금까지 응답해 주셔서 대단히 감사합니다. 좋은 하루 되세요. ▣</div>
(읽어주고 조사 종료) 본 조사에 대한 문의는 02-2122-%s로 연락 주시기 바랍니다.''',
        'KSA'  : '''<div align=center style='border: 1px solid #800000;padding:10px; white-space: normal; font-size:13px; border-collapse: collapse;background-color:#5579d7;color:#ffffff'>
▣ 아래 내용 읽어주고 조사 종료 ▣</div>
<div style="text-align:center;">
지금까지 기획재정부에서 주관하는 공공기관 고객만족도 조사 공동실사단의 면접원 OOO이었습니다.<br/>
본 조사에 대한 문의는 02-2122-%s로 연락 주시기 바랍니다.<br/>
응답 해 주셔서 대단히 감사합니다. 좋은 하루 되세요.
</div>'''
    }

    # 마무리 인사 및 실사 담당자 세팅
    for r, c in [ (75, 4), (76, 4) ] :
      last_page = new_ws.cell(r, c)
      last_page.value = last_page.value%(last_text[division]%da_num[da])

    # 설문지 분류 셀 관련
    cell_value = []
    for code, attr in change_cells['qnrs'].items() :
        qname = attr['name']
        word = f'({code}) {qname}'
        cell_value.append(word)


    set_cells = [(13, 6), (25, 6)]
    for r, c in set_cells :
        set_sell = new_ws.cell(r, c)
        set_sell.value = set_sell.value%('\n'.join(cell_value))

    # 기관명 & 설문 분류
    js_logics = []
    for code, attr in change_cells['qnrs'].items() :
        qname = attr['name']
        word = f'if(ADD05=={code}){{text="{qname}";}}'
        js_logics.append(word)

    QQQ1_set = new_ws.cell(16, 8)
    QQQ1_set.value = QQQ1_set.value%(survey_name, '\n'.join(js_logics))

    # 설문 타입 오토펀치 syntax
    q_type_quto = []
    for code, attr in change_cells['qnrs'].items() :
        qtype = attr['type_code']
        word = f'if(QQQ14=={code}) then TQ1={qtype}'
        q_type_quto.append(word)

    TQ1_set = new_ws.cell(29, 7)
    TQ1_set.value = TQ1_set.value%('\n'.join(q_type_quto))

    # 워딩 다른 문항 출력
    Q_cell_dict = {
        'SQ3' : (30, 8),
        'SQ4' : (31, 8),
        '문 9' : (61, 8),
        '문 10' : (62, 8), 
        '문 12' : (64, 8),
    }

    for qid in key_texts :
        cr, cc = Q_cell_dict[qid]
        curr_cell = new_ws.cell(cr, cc)
        js_logics = []
        for code, txt in change_cells[qid].items() :
            word = f'if(QQQ14=={code}){{text="{txt}";}}'
            js_logics.append(word)
        
        curr_cell.value = curr_cell.value%('\n'.join(js_logics))


    # SQ 문항 이후 안내 문구 출력
    info_texts = []
    for code, txt in change_cells['info'].items() :
        word = f'if(QQQ14=={code}){{text="{txt}";}}'
        info_texts.append(word)



    # QQQ1/Q1/Q5/Q6 워딩 구분
    QQQ1_txt = {
        'KMAC' : '2023년 1~2월 기획재정부 공동실사단',
        'KSA'  : '2023년 1~2월 기획재정부/닐슨아이큐코리아(유)'    
    }
    Q1_txt = {
        'KMAC' : '각 질문에 대해 고객님께서 동의하시는 정도에 따라 보기(11개) 중에서 골라주세요.<br/>0점(전혀 그렇지 않다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 그렇다)까지 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.',
        # 'KSA'  : '각 질문에 대해 고객님께서 동의하시는 정도에 따라 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점 중에서 골라주세요. 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.'
        'KSA'  : '각 질문에 대해 고객님께서 동의하시는 정도에 따라 보기(11개) 중에서 골라주세요.<br/>0점(전혀 그렇지 않다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 그렇다)까지 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.'
    }


    Q5_txt = {
        'KMAC' : '<br/>0점(매우 나쁘다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 좋다)까지 긍정적일수록 높은 점수, 부정적 일수록 낮은 점수를 주시면 됩니다.',
        #'KSA'  : '각 질문에 대해 고객님께서 동의하시는 정도에 따라 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점 중에서 골라주세요. 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.'
        'KSA' : '<br/>0점(매우 나쁘다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 좋다)까지 긍정적일수록 높은 점수, 부정적 일수록 낮은 점수를 주시면 됩니다.'
    }


    Q6_txt = {
        'KMAC' : '<br/>0점(전혀 그렇지 않다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 그렇다)까지 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.',
        # 'KSA'  : '각 질문에 대해 고객님께서 동의하시는 정도에 따라 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점 중에서 골라주세요. 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.'
        'KSA'  : '<br/>0점(전혀 그렇지 않다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 그렇다)까지 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.'
    }

    QQQ1_cell = new_ws.cell(16, 4)
    QQQ1_cell.value = QQQ1_cell.value.format(word=QQQ1_txt[division])

    Q1_cell = new_ws.cell(35, 4)
    Q1_cell.value = Q1_cell.value.format(word=Q1_txt[division])

    Q5_cell = new_ws.cell(46, 4)
    Q5_cell.value = Q5_cell.value.format(word=Q5_txt[division])

    Q6_cell = new_ws.cell(51, 4)
    Q6_cell.value = Q6_cell.value.format(word=Q6_txt[division])

    # KSA Q8 세팅
    if division == 'KSA' :
        max_array_length = max([len(i) for i in change_cells['Q8'].values()])

        # Q8 row setting
        Q8X1_qid_cell = new_ws.cell(58, 3)
        Q8X1_qid_cell.value = Q8X1_qid_cell.value%(max_array_length)

        # exmpample setting
        Q8X1_ex = [f'({i}) <span class=\'qnum\'>{i}</span>. <span class=\'row_text\'></span>' for i in range(1, max_array_length+1)]
        Q8X1_ex_cell = new_ws.cell(58, 6)
        Q8X1_ex_cell.value = Q8X1_ex_cell.value%('\n'.join(Q8X1_ex))

        Q8X2_ex = [f'({i}) {i}. <span class=\'SR_row_text\'></span>' for i in range(1, max_array_length+1)]
        Q8X2_ex_cell = new_ws.cell(59, 6)
        Q8X2_ex_cell.value = Q8X2_ex_cell.value%('\n'.join(Q8X2_ex))

        # js setting
        Q8_array = [f'if( QQQ14=={code} ){{ textArr = {arr};}}' for code, arr in change_cells['Q8'].items()]
        Q8X1_js_cell = new_ws.cell(57, 8)
        Q8X3_js_cell = new_ws.cell(60, 8)

        Q8X1_js_cell.value = Q8X1_js_cell.value%('\n'.join(Q8_array))
        Q8X3_js_cell.value = Q8X3_js_cell.value%('\n'.join(Q8_array))

    # CC 세팅
    cc_1_cell = new_ws.cell(74, 7)
    # if division == 'KMAC' :
    #   new_ws.cell(73, 7).value = None
    #   cc_1_cell.value = cc_1_cell.value%('')

    if division == 'KSA' :
      cc_1_cell.value = cc_1_cell.value%('display_yn(n)')

    # 기관별 문항 타입 구분
    if division == 'KMAC' :
        # DQ2
        new_ws.cell(66, 7).value = None
        new_ws.delete_rows(68, 3)
        new_ws.delete_rows(56, 9)

    if division == 'KSA' :
        # DQ2X1, DQ2X2
        new_ws.delete_rows(67, 1)
        new_ws.delete_rows(55, 1)

    info_cell = new_ws.cell(33, 8)
    info_cell.value = info_cell.value%('\n'.join(info_texts))

    # save
    save_filename = f'PCSI_{division}_{survey_name}.xlsx'
    new_wb.save(os.path.join(os.getcwd(), save_folder, save_filename))


    print('💠 PCSI 스마트 서베이 확인 사항')
    print('   - xls로 다시 저장할 것')
    print('   - SQ/DQ 밑 설문지별 수정되는 변수 확인 필요')
    print('   - SQ1/SQ2도 설문지에 따라 다를 수 있음')
    print('   - DQ2 문항 : KMAC은 개인/법인 상관없이 DQ2에서 직업만 확인')
    print('   - DQ2 문항 : KSA는 개인의 경우 직업, 법인의 경우 직원수를 질문')
    print('   - 실사 담당자 전화번호 확인')
    print('   - 실사 시작전에 히든 변수 display_yn(n) 설정 해줄 것')
    print('   - 쿼터 세팅 확인')

    # 현장조사세팅
    # KQ Autopunch
    rnum = 1 if division == 'KMAC' else 2 if division == 'KSA' else ''
    KQ_auto = f'KQ.val = KQ.r{rnum}.index'

    # 설문분류 rows
    xml_QQQ14 = []
    for code, attr in change_cells['qnrs'].items() :
        qname = attr['name']
        word = f'<row label="r{code}" value="{code}">{qname}</row>'
        xml_QQQ14.append(word)

    xml_QQQ14 = '\n'.join(xml_QQQ14)

    # 설문 타입 오토펀치 exec
    xml_type_auto = []
    for code, attr in change_cells['qnrs'].items() :
        qtype = attr['type']
        word = f'if QQQ14.r{code} : TQ1.val = TQ1.{qtype}.index'
        xml_type_auto.append(word)

    xml_type_auto = '\n'.join(xml_type_auto)

    # pipe case setting
    xml_q_case = {
        'SQ3' : '',
        'SQ4' : '',
        '문 9' : '',
        '문 10' : '',
        '문 12' : '',
    }

    for qid in key_texts :
        case = []
        for code, txt in change_cells[qid].items() :
            curr_txt = txt.replace('<font color=blue>', '<strong>')
            curr_txt = curr_txt.replace('</font>', '</strong>')
            curr_txt = curr_txt.replace('<', '&lt;')
            curr_txt = curr_txt.replace('>', '&gt;')
            word = f'<case label="r{code}" cond="QQQ14.r{code}">{curr_txt}</case>'
            case.append(word)
        
        xml_q_case[qid] = '\n'.join(case)

    # SQ 문항 이후 INFO
    xml_info = []
    for code, txt in change_cells['info'].items() :
        curr_txt = txt.replace('<font color=blue>', '<strong>')
        curr_txt = curr_txt.replace('</font>', '</strong>')
        curr_txt = curr_txt.replace('<', '&lt;')
        curr_txt = curr_txt.replace('>', '&gt;')
        word = f'<case label="r{code}" cond="QQQ14.r{code}">{curr_txt}</case>'
        xml_info.append(word)

    xml_info = '\n'.join(xml_info)


    # KSA Q8 XML 세팅
    Q8_array = ''
    Q8_rows = ''
    HQ8_rows = ''

    if division == 'KSA' :
        # exec setting
        Q8_array = {f'r{code}' : arr for code, arr in change_cells['Q8'].items()}
        
        # rows setting
        max_array_length = max([len(i) for i in change_cells['Q8'].values()])

        HQ8_rows = []
        Q8_rows = []
        
        for idx, i in enumerate(list(range(1, max_array_length+1))) :
            hword = f'<row label="r{i}" value="{i}"/>'
            HQ8_rows.append(hword)
            
            word = f'<row label="_{i}" value="{i}">{i}. ${{HQ8X1.r{i}.unsafe_val}}</row>'
            Q8_rows.append(word)

        Q8_rows = '\n'.join(Q8_rows)
        HQ8_rows = '\n'.join(HQ8_rows)


    op_num = ''
    if division == 'KMAC' :
      op_num = 'OP-000141831'
    if division == 'KSA' :
      op_num = 'OP-000155091'

    xml_Q8_after = {
        'KMAC' : f'''<textarea
  label="Q8"
  optional="0"
  width="100">
  <title>문 8】 마지막으로 <strong>${{res.pcsi_name}}</strong>에 바라시는 점이 있다면 자유롭게 말씀해 주십시오.</title>
  <comment></comment>
</textarea>
<suspend/>''',

        'KSA' : f'''<text
  label="HQ8X1"
  size="40"
  optional="1"
  where="execute">
  <title>(HIDDEN) Q8-1 제시 속성</title>
  <comment></comment>
<exec>
for eachRow in HQ8X1.rows :
	eachRow.val = None

Q8attrs = {Q8_array}
attrs = Q8attrs[QQQ14.selected.label]
for idx, attr in enumerate(attrs) :
  HQ8X1.rows[idx].val = attr
</exec>
{HQ8_rows}
</text>
<suspend/>


<define label="Q8X1_list">
{Q8_rows}
</define>

<suspend/>

<radio
  label="Q8X1"
  rowCond="HQ8X1.rows[row.index]"
  surveyDisplay="desktop">
  <title>문 8】 <strong>${{res.pcsi_name}}</strong>에서 경험하신 서비스의 단계에 대해 고객님께서 만족하시는 정도에 따라 0점(매우 불만족), 1, 2, 3, 4, 5, 6, 7, 8, 9, 10(매우 만족)점 중에서 골라주세요.
만족하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>매우<br/>불만족</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>이다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>매우<br/>만족</col>
  <insert source="Q8X1_list" />
</radio>
<suspend/>

<checkbox
  label="HQ8X2"
  atleast="0"
  rowCond="HQ8X1.rows[row.index]"
  where="execute">
  <title>(HIDDEN) Q8X1 낮은 점수</title>
  <exec>
Q8X1_answers = [eachRow.val for eachRow in Q8X1.rows if eachRow.displayed]
min_value = min(Q8X1_answers)
for eachRow in HQ8X2.rows :
  eachRow.val = False
  if eachRow.displayed :
    if Q8X1.attr(eachRow.label).val == min_value :
      eachRow.val = True
  </exec>
  <comment></comment>
  <insert source="Q8X1_list" />
</checkbox>
<suspend/>


<radio
  label="Q8X2"
  rowCond="HQ8X2[row]">
  <title>문 8-2】 그중에서 가장 불만족한 단계는 무엇입니까?</title>
  <comment></comment>
  <insert source="Q8X1_list" />
</radio>
<suspend/>

<textarea
  label="Q8X3"
  optional="0"
  width="100">
  <title>문 8-3】 그렇다면 위에서 가장 만족도가 낮은 <strong>[pipe: Q8X2] 단계</strong>에서 가장 불편하거나 불만스러웠던 점은 무엇이었습니까?</title>
  <comment></comment>
</textarea>
<suspend/>

<pipe
  label="Q9_pipe"
  capture="">
{xml_q_case['문 9']}
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="Q9">
  <title>문 9】 [pipe: Q9_pipe]</title>
  <comment></comment>
  <row label="r1" value="1">1년 미만</row>
  <row label="r2" value="2">1년 이상 ~ 3년 미만</row>
  <row label="r3" value="3">3년 이상 ~ 5년 미만</row>
  <row label="r4" value="4">5년 이상</row>
</radio>
<suspend/>

<pipe
  label="Q10_pipe"
  capture="">
{xml_q_case['문 10']}
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="Q10">
  <title>문 10】 [pipe: Q10_pipe]</title>
  <comment></comment>
  <row label="r1" value="1">스스로 검색 및 탐색</row>
  <row label="r2" value="2">기관의 안내(공문, 이메일 등)</row>
  <row label="r3" value="3">주위 지인들의 추천</row>
  <row label="r4" value="4">광고/홍보(신문광고 등)</row>
  <row label="r5" value="5" open="1" openSize="25" randomize="0">기타(적을 것 :)</row>
</radio>
<suspend/>


<radio 
  label="Q11">
  <title>문 11】 고객님께는 전년도에 <strong>${{res.pcsi_name}}</strong>의 고객만족도 설문조사에 참여하셨습니까?</title>
  <row label="r1" value="1">예</row>
  <row label="r2" value="2">아니오</row>
</radio>
<suspend/>


<pipe
  label="Q12_pipe"
  capture="">
{xml_q_case['문 12']}
  <case label="null" cond="1">UNDEFINED</case>
</pipe>


<textarea
  label="Q12"
  optional="0"
  width="100">
  <title>문 12】 [pipe: Q12_pipe]</title>
  <comment></comment>
</textarea>
<suspend/>
        '''
    }

    xml_DQ2 = {
        'KMAC' : f'''
  <radio 
   label="DQ2"
   randomize="0">
    <title>DQ2. 고객님의 직업은 어떻게 되십니까? </title>
    <row label="r1" value="1">전문직</row>
    <row label="r2" value="2">경영직</row>
    <row label="r3" value="3">관리직</row>
    <row label="r4" value="4">사무직</row>
    <row label="r5" value="5">생산직</row>
    <row label="r6" value="6">영업직</row>
    <row label="r7" value="7">노무직</row>
    <row label="r8" value="8">판매/서비스직</row>
    <row label="r9" value="9">자영업</row>
    <row label="r10" value="10">농/임/축산업</row>
    <row label="r11" value="11">전업주부</row>
    <row label="r12" value="12">학생</row>
    <row label="r13" value="13">무직</row>
    <row label="r14" randomize="0" value="14">기타</row>
  </radio>
  <suspend/>
        ''',

        'KSA' : f'''<radio
  label="DQ2X1"
  cond="QQQ12.r1">
  <title>DQ2. 고객님의 직업은 어떻게 되십니까?</title>
  <comment></comment>
  <row label="r1" value="1">전문직</row>
  <row label="r2" value="2">경영직</row>
  <row label="r3" value="3">관리직</row>
  <row label="r4" value="4">사무직</row>
  <row label="r5" value="5">생산직</row>
  <row label="r6" value="6">영업직</row>
  <row label="r7" value="7">노무직</row>
  <row label="r8" value="8">판매/서비스직</row>
  <row label="r9" value="9">자영업</row>
  <row label="r10" value="10">농/임/축산업</row>
  <row label="r11" value="11">전업주부</row>
  <row label="r12" value="12">학생</row>
  <row label="r13" value="13">무직</row>
  <row label="r14" randomize="0" value="14">기타</row>
</radio>
<suspend/>


<number
  label="DQ2X2"
  size="3"
  cond="QQQ12.r2"
  optional="0"
  ss:postText="명"
  verify="range(0, 99999)">
  <title>DQ2. 고객님께서 속해계시는 사업체의 직원 수는 몇 명입니까?</title>
  <comment></comment>
  <noanswer label="na">해당 없음</noanswer>
</number>
<suspend/>

<radio
  label="DQ2X2_1"
  cond="QQQ12.r2"
  where="execute">
  <title>(HIDDEN) DQ2 AUTO</title>
  <comment></comment>
  <exec>
if DQ2X2.na :
  DQ2X2_1.val = DQ2X2_1.r8.index
else :
  persons = ['&lt;50', '50-99', '100-199', '200-299', '300-599', '500-999', '&gt;=1000']

  for idx, person in enumerate(persons) :
    if DQ2X2.check(person) :
      DQ2X2_1.val = idx
      break
  </exec>
  <row label="r1" value="1">50인 미만</row>
  <row label="r2" value="2">50인~100인 미만</row>
  <row label="r3" value="3">100인~200인 미만</row>
  <row label="r4" value="4">200인~300인 미만</row>
  <row label="r5" value="5">300인~500인 미만</row>
  <row label="r6" value="6">500인~1000인 미만</row>
  <row label="r7" value="7">1,000인 이상</row>
  <row label="r8" value="8">해당 없음</row>
</radio>
<suspend/>
        '''
    }


    xml = f'''<?xml version="1.0" encoding="UTF-8"?>
<survey 
  alt="{op_num}_PCSI_{division}_{survey_name}_현장조사"
  autosaveKey="UID"
  browserDupes=""
  builder:wizardCompleted="1"
  builderCompatible="1"
  compat="152"
  delphi="1"
  displayOnError="all"
  extraVariables="ipAddress,source,record,decLang,list,userAgent"
  fir="on"
  html:showNumber="0"
  lang="korean"
  markerTimeout="0"
  mobile="compat"
  mobileDevices="smartphone,tablet,desktop"
  name="Survey"
  persistentExit="1"
  secure="1"
  setup="term,decLang,quota,time"
  ss:disableBackButton="1"
  ss:enableNavigation="1"
  ss:hideProgressBar="0"
  ss:includeCSS="/survey/selfserve/nielseniq.css"
  state="testing"
  theme="company/nielseniq-new">

<res label="chk_plz">응답 확인 부탁드립니다.</res>
<res label="samegrid">모든 항목에 대해 동일한 답변을 입력했습니다.</res>
<res label="softerr">응답을 다시 한번 확인해 주세요. 응답이 맞을 경우, 다음버튼을 누르고 진행을 하시면 됩니다.</res>
<res label="err2010">하나 이상의 답변을 입력하십시오.</res>
<res label="err2011">동일한 답변을 입력하지 마십시오.</res>
<res label="err2012">첫 번째 텍스트 입력란부터 순서대로 입력하십시오.</res>
<res label="err2050">중복 입력하지 마십시오.</res>
<res label="badhan">ㄱ,ㄲ,ㄴ,ㄷ,ㄸ,ㄹ,ㅁ,ㅂ,ㅃ,ㅅ,ㅆ,ㅇ,ㅈ,ㅉ,ㅊ,ㅋ,ㅌ,ㅍ,ㅎ,ㅏ,ㅐ,ㅑ,ㅒ,ㅓ,ㅔ,ㅕ,ㅖ,ㅗ,ㅘ,ㅙ,ㅚ,ㅛ,ㅜ,ㅝ,ㅞ,ㅟ,ㅠ,ㅡ,ㅢ,ㅣ,ㄵ,ㄶ,ㄺ,ㄻ,ㄼ,ㄽ,ㄾ,ㄿ,ㅀ,ㅄ</res>
<res label="badhan_err">입력 확인 부탁 드립니다.</res>
<res label="badspa">@,$,%,#,*,!,?</res>
<res label="badspa_err">특수 문자는 입력하실 수 없습니다.</res>
<res label="block_ie">Internet Explorer는 지원하지 않습니다.</res>
<res label="cardrating_msg">'계속' 버튼을 눌러 다음 질문에 응답해주세요.</res>
<samplesources default="0">
  <samplesource list="0">
    <title>Open Survey</title>
    <completed>귀하께서는 이미 본 설문에 접속 하신 것으로 확인 됩니다.</completed>
    <exit cond="terminated"><strong>본 조사에 참여해주셔서 감사합니다.</strong><br /><br />본격적으로 조사를 시작하기 전에, 귀하가 본 조사에 적합한 응답 대상인지 알아보기 위해 몇 가지의 질문을 드렸습니다.<br /><br />죄송합니다. 귀하께서는 본 조사의 응답 대상이 아니십니다.<br /><br />차후에 다른 온라인 조사에 참여해주시면 감사하겠습니다.<br /><br />귀하의 소중한 의견은 더 나은 제품과 서비스를 개발하는데 좋은 정보가 될 것입니다.</exit>
    <exit cond="qualified">▣ 지금까지 응답해 주셔서 대단히 감사합니다. 좋은 하루 되세요. ▣
  <br /><br />※ 본 조사에 대한 문의는 아래의 연락처로 주시기 바랍니다.
 <br />- 연구 책임자 : 김진규 (☎ 02-2122-7357)
 <br />- 실사 책임자 : {da} (☎ 02-2122-{da_num[da]})</exit>
    <exit cond="overquota"><strong>본 조사에 참여해 주셔서 감사합니다.</strong><br /><br />안타깝게도, 귀하께서 해당하시는 조사 대상 그룹의 조사는 이미 종료되었습니다.<br /><br />다음에 참여해 주시기 바랍니다.</exit>
  </samplesource>

  <samplesource list="9">
    <title>UserIdSampleSource</title>
    <invalid>URL에 정보가 누락되었습니다. 기존의 초대받은 URL을 확인해 주시기 바랍니다.</invalid>
    <completed>귀하께서는 이미 본 설문에 접속 하신 것으로 확인 됩니다.</completed>
    <var name="UID" unique="1"/>
    <exit cond="terminated"><strong>본 조사에 참여해주셔서 감사합니다.</strong><br /><br />본격적으로 조사를 시작하기 전에, 귀하가 본 조사에 적합한 응답 대상인지 알아보기 위해 몇 가지의 질문을 드렸습니다.<br /><br />죄송합니다. 귀하께서는 본 조사의 응답 대상이 아니십니다.<br /><br />차후에 다른 온라인 조사에 참여해주시면 감사하겠습니다.<br /><br />귀하의 소중한 의견은 더 나은 제품과 서비스를 개발하는데 좋은 정보가 될 것입니다.</exit>
    <exit cond="qualified"><strong>이로써 설문이 완료되었습니다.</strong><br /><br /><strong>귀한 시간 내주셔서 대단히 감사드립니다.</strong></exit>
    <exit cond="overquota"><strong>본 조사에 참여해 주셔서 감사합니다.</strong><br /><br />안타깝게도, 귀하께서 해당하시는 조사 대상 그룹의 조사는 이미 종료되었습니다.<br /><br />다음에 참여해 주시기 바랍니다.</exit>
  </samplesource>
</samplesources>

<suspend/>

<style name="respview.client.css"><![CDATA[
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/animate.css/4.1.1/animate.min.css"/>
<link rel="stylesheet" href="[rel util.css]"/>
<style>
.hangle-span{{color:#e7046f;border:0;text-align:right;width:200px; }}
.han-input-div{{display: flex; gap: 10px;}}

\@media (prefers-color-scheme: dark) {{
  .non-touch .fir-icon:hover .rounded .fir-base, .non-touch .fir-icon:hover .square .fir-base, .non-touch .grid-table-mode .clickableCell:hover .fir-icon .rounded .fir-base, .non-touch .grid-table-mode .clickableCell:hover .fir-icon .square .fir-base {{ fill: #878787; }}
  .fir-icon.selected .rounded .fir-selected, .fir-icon.selected .square .fir-selected {{ fill: #878787; }}
}}

.sq-cardsort-bucket-count{{display: none !important;}}

.sq-cardrating-content{{
  font-size: 1.2rem !important;
}}
\@media all and (min-width: 770px){{
 .sq-cardrating-content{{
  font-size: 1.4rem !important;
 }}
 
  .han-input-div{{
    display: block;
    gap: 0px;
  }}
}}

.sq-sliderpoints-container{{
  max-width: 900px;
}}

.sq-sliderpoints-container{{
  margin-left: 0 !important;
}}

.dq-imgupload{{
  max-width: 924px;
}}

.sq-accordion-row{{
  max-width: 924px !important;
  font-size: 1.2rem;
}}

.sq-cardsort{{
  margin: 0 !important;
  max-width: 924px;
}}

.input-max-width input[type=number]{{
  max-width: 70px;
}}

[data-viewsize=small] [data-viewmode=horizontal] .sq-cardrating-button{{
  max-height: 70px;
}}

.gridRank .grid{{
    display: none !important;
}}
.survey-body{{
  margin-top: 5% !important;
}}
</style>
]]></style>
<style name="respview.client.js"><![CDATA[
<script src="[rel mouse.js]"></script>
<script src="[rel util.js]"></script>
<script src="[rel animation.js]"></script>
<script src="[rel uses.js]"></script>
<script>
/** DOM Loaded Animation function **/
// const animation_start = 'left';
// const animation_distance = 50;
// const animation_speed = 0.3;
// const animation_delay = 0.05;
// fnInitAnimation(animation_start, animation_distance, animation_speed, animation_delay);

window.addEventListener("DOMContentLoaded", function(){{
  /** Restart remove **/
  const restart = document.querySelector(".autosave-restart");
  if( restart ){{ restart.remove(); }}

  if (!document.querySelector('.controlbarContainer')){{
    document.oncontextmenu = function(){{ return false; }}  /** mouse right click **/
    document.onselectstart = function(){{ return false; }}  /** mouse text drag block **/
    document.ondragstart = function(){{ return false; }}    /** mouse image drag block **/
  }}

  /** print screen key block **/
  document.addEventListener('keyup', function(e){{
      if (e.key == 'PrintScreen') {{
          navigator.clipboard.writeText('');
      }}
  }});
  document.addEventListener('keydown', function(e){{
      if (e.ctrlKey && e.key == 'p') {{
          e.cancelBubble = true;
          e.preventDefault();
          e.stopImmediatePropagation();
      }}
  }});

  /** rank question setting (grid type) **/
  gridRankSetting("gridRank", "exclusive");
  
  /** accordian show **/
  accordionFirstRowShow();
  accordionErrorHighLight();
}});
</script>
]]></style>
<suspend/>

<style label="hideElements" name="question.element"><![CDATA[
\@if ec.simpleList
<div class="element $(rowStyle) $(levels) $(extraClasses) ${{col.group.styles.ss.groupClassNames if col.group else (row.group.styles.ss.groupClassNames if row.group else "")}} $(col.styles.ss.colClassNames) $(row.styles.ss.rowClassNames) ${{"clickableCell" if isClickable else ""}}"$(extra)>
    ${{v2_insertStyle('el.label.start')}}
    ${{v2_insertStyle('el.label.end')}}
</div>
\@else
<$(tag) $(headers) class="cell nonempty element $(levels) ${{"desktop" if this.grouping.cols else "mobile"}} border-collapse $(extraClasses) ${{col.group.styles.ss.groupClassNames if col.group else (row.group.styles.ss.groupClassNames if row.group else "")}} $(col.styles.ss.colClassNames) $(row.styles.ss.rowClassNames) ${{"clickableCell" if isClickable else ""}}"$(extra)>
    ${{v2_insertStyle('el.label.start')}}
    ${{v2_insertStyle('el.label.end')}}
</$(tag)>
\@endif
]]></style>

<style cond="1" name="survey.respview.footer.support"><![CDATA[
<div></div>
]]></style>

<suspend/>

<exec when="init">
survey_path = gv.survey.path.split("/")[-2] if gv.survey.path.split("/")[-1] in ["temp-edit-live", "temp-view"] else gv.survey.path.split("/")[-1]
# Adhoc = AD / Tracking = TRC
imgdr = "https://nielsenkor.cafe24.com/Decipher/AD/{{}}".format(survey_path)

from datetime import datetime
import random

def status(condt,label):
    if condt : 
      RespStatus.val=getattr(RespStatus,"{{}}".format(label)).index

def soft_Err(cond, str, high_rows=[], high_cols=[]):
  if cond :
    if p.chk==0 :
      p.chk=1
    else :
      p.chk=0
      if high_rows :
        for each in high_rows :
          error(str, row=each)

      elif high_cols :
        for each in high_cols :
          error(str, col=each)
      else :
        error(str)
  else :
    p.chk=1

def badtext() :
  tt=[i.replace(" ","") for i in this.values if not i==None]
  badhan=(res.badhan)
  badhan=badhan.split(',')

  badspa=(res.badspa)
  badspa=badspa.split(',')


  for i in range(0,len(tt)) :
    current_v=tt[i]
    current_len=len(current_v)
    for j in range(0,int(current_len),3) :
      fn=j
      ln=int(j)+3
      if current_v[fn:ln] in badhan :
          error(res.badhan_err)


  for i in range(0,len(tt)) :
    current_v=tt[i]
    current_len=len(current_v)
    for j in range(0,int(current_len),1) :
      if current_v[j] in badspa :
        error(res.badspa_err)
</exec>

<suspend/>

<exec>
p.chk = 1
</exec>

<suspend/>


<quota label="tot" overquota="noqual" sheet="tot"/>

<suspend/>

<radio 
  label="RespStatus"
  where="execute,survey,report">
  <title>STATUS</title>
  <exec>
#incomplete
status(True,'r2')
  </exec>

  <row label="r1" value="1">complete</row>
  <row label="r2" value="2">incomplete</row>
  <row label="AgreeC" value="3">Agree-Close</row>
  <row label="Scr_Video" value="4">Scr_Video</row>
  <row label="Scr_Audio" value="5">Scr_Audio</row>
  <row label="SQ1" value="101">SQ1 ScreenOut</row>
  <row label="SQ2" value="102">SQ2 ScreenOut</row>
  <row label="SQ3" value="103">SQ3 ScreenOut</row>
  <row label="r96" value="96">Removed - Over Quota/ Sample</row>
  <row label="r97" value="97">Removed - QC</row>
  <row label="r98" value="98">Unsubscribed</row>
  <row label="r99" value="99">Failed Data Trapping Test</row>
</radio>

<suspend/>

<res label="pcsi_name">{survey_name}</res>


<radio
  label="KQ"
  where="execute">
  <title>(HIDDEN) 구분</title>
  <comment></comment>
  <exec>
{KQ_auto}
  </exec>
  <row label="r1" value="1">KMAC</row>
  <row label="r2" value="2">KSA</row>
</radio>

<suspend/>


<radio
  label="QQQ12">
  <title>개인/법인 구분</title>
  <comment></comment>
  <row label="r1" value="1">개인</row>
  <row label="r2" value="2">법인</row>
</radio>

<radio
  label="QQQ14">
  <title><strong>설문 분류</strong></title>
  <comment></comment>
{xml_QQQ14}
</radio>

<suspend/>

<style name="survey.completion"><![CDATA[
\@if not gv.survey.root.styles.ss.hideProgressBar
    <div class="progress-bar progress-${{"top" if gv.survey.root.progressOnTop else "bottom"}}" title="@(progress-bar) - $(percent)% @(complete)">
      <div class="progress-box-outer"><span class="progress-box-completed" style="width: $(percent)%;"></span></div>
      <div class="progress-text"><span class="screen-readers-only">@(progress-bar) </span>$(percent)%</div>
    </div>
    <div><strong>${{'- %s'%(QQQ14.selected.text) if QQQ14.any else ''}}</strong></div>
    <div><strong>${{'- %s'%(QQQ12.selected.text) if QQQ12.any else ''}}</strong></div>
\@endif
]]></style>

<suspend/>

<text 
  label="InterData"
  optional="0"
  size="40">
  <title>면접원 정보</title>
<style name='el.text' rows="time"> <![CDATA[
\@if row.styles.ss.preText or this.styles.ss.preText
    ${{row.styles.ss.preText or this.styles.ss.preText or ""}}&nbsp;
\@endif
<select name="$(name)" class="input dropdown"> 
<option></option>
\@for item in range(9, 19)
<option value="$(item)" ${{"SELECTED" if str(item)==ec.value else ""}}>$(item)</option> 
\@end
</select>
\@if row.styles.ss.postText or this.styles.ss.postText
    &nbsp;${{row.styles.ss.postText or this.styles.ss.postText or ""}}
\@endif
]]></style>
  <row label="name" ss:preText="이름" size="8"/>
  <row label="time" ss:preText="시작 시간" ss:postText="시"/>
  <row label="area" ss:preText="조사 장소"/>
</text>

<suspend/>

<pipe
  label="QQQ1_pipe"
  capture="">
  <case label="r1" cond="KQ.r1">2023년 1~2월 기획재정부 공동실사단</case>
  <case label="r2" cond="KQ.r2">2023년 1~2월 기획재정부</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<html label="QQQ1">
안녕하십니까?<br/>
저는 기획재정부에서 주관하는 「공공기관 고객만족도 조사 : <strong>${{res.pcsi_name}}</strong>」을 담당한 면접원 ○○○입니다.<br/><br/>
본 조사는「공공기관의 운영에 관한 법률」 제13조에 의거하여 실시하고 있으며, 응답하신 내용은 통계를 내는 데만 사용될 뿐, 외부에 노출되지 않으며, 통계법에 의거, 절대 비밀이 보장됩니다.<br/>
바쁘시겠지만 공공기관의 서비스 개선을 위해 응답해 주시면 대단히 감사하겠습니다.<br/><br/>[pipe: QQQ1_pipe]
</html>

<suspend/>

<radio
  label="TQ1"
  where="execute">
  <title>(HIDDEN) 설문지 타입</title>
  <comment></comment>
  <exec>
{xml_type_auto}
  </exec>
  <row label="A" value="1">A</row>
  <row label="B" value="2">B</row>
  <row label="C" value="3">C</row>
  <row label="D" value="4">D</row>
  <row label="E" value="5">E</row>
  <row label="F" value="6">F</row>
  <row label="G" value="7">G</row>
  <row label="H" value="8">H</row>
</radio>
<suspend/>

<radio
  label="SQ1">
  <title>SQ1. 혹시 고객님의 가족이나 친지 중에 다음 업종에 종사하고 계신 분이 있으신지요?</title>
  <comment></comment>
  <row label="r1" value="1"><strong>${{res.pcsi_name}}</strong> 직원</row>
  <row label="r2" value="2">광고회사나 시장조사 회사</row>
  <row label="r99" value="99" randomize="0">없다</row>
</radio>
<suspend/>

<exec>
status(not SQ1.r99,"SQ1")
</exec>

<term label="SQ1_Term" cond="not SQ1.r99">SQ1 term</term>

<suspend/>


<number
  label="SQ2"
  size="3"
  optional="0"
  verify="range(0,100)"
  ss:questionClassNames="input-max-width"
  ss:preText="만"
  ss:postText="세">
  <title>SQ2. 실례지만 고객님의 ‘연세’는 만으로 어떻게 되시나요?</title>
  <comment></comment>
</number>
<suspend/>

<note>설문별 Range 확인</note>
<exec>
status(not SQ2.check('20-65'),"SQ2")
</exec>

<term label="SQ2_Term" cond="not SQ2.check('20-65')">SQ2 term</term>

<suspend/>

<pipe
  label="SQ3_pipe"
  capture="">
{xml_q_case['SQ3']}
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="SQ3">
  <title>SQ3. [pipe: SQ3_pipe]</title>
  <comment></comment>
  <row label="r1" value="1">예</row>
  <row label="r2" value="2">아니오(잘 모르겠다)</row>
</radio>
<suspend/>


<exec>
status(SQ3.r2,"SQ3")
</exec>

<term label="SQ3_Term" cond="SQ3.r2">SQ3 term</term>

<suspend/>


<pipe
  label="SQ4_pipe"
  capture="">
{xml_q_case['SQ4']}
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="SQ4">
  <title>SQ4. [pipe: SQ4_pipe]</title>
  <comment></comment>
  <row label="r1" value="1">예</row>
  <row label="r2" value="2">아니오</row>
</radio>
<suspend/>

<pipe
  label="INFO1_pipe"
  capture="">
{xml_info}
  <case label="null" cond="1">UNDEFINED</case>
</pipe>


<html label="INFO1" cond="SQ4.r1">
[pipe: INFO1_pipe]
</html>

<suspend/>

<pipe
  label="Q1_pipe"
  capture="">
  <case label="r1" cond="KQ.r1">각 질문에 대해 고객님께서 동의하시는 정도에 따라 보기(11개) 중에서 골라주세요.<br/>0점(전혀 그렇지 않다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 그렇다)까지 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.</case>
  <case label="r2" cond="KQ.r2">각 질문에 대해 고객님께서 동의하시는 정도에 따라 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점 중에서 골라주세요. 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>


<pipe
  label="Q1_row1_pipe"
  capture="">
  <case label="r1" cond="TQ1.selected.label in ['A', 'B']">을 통해 이용한 시설, 장비, 제품 등은 나의 이용 목적에 부합한다.</case>
  <case label="r2" cond="TQ1.selected.label in ['C', 'D']">을 통해 이용한 서비스는 나의 이용 목적에 부합한다.</case>
  <case label="r3" cond="TQ1.selected.label in ['E', 'F']">을 통해 이용한 시설, 장비, 제품 등은 기관의 사업 목적에 부합한다.</case>
  <case label="r4" cond="TQ1.selected.label in ['G', 'H']">을 통해 이용한 서비스는 기관의 사업 목적에 부합한다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<pipe
  label="Q1_row2_pipe"
  capture="">
  <case label="r1" cond="TQ1.selected.label in ['A', 'E']">의 시설, 장비, 제품 등을 통해 목적을 달성할 수 있다.</case>
  <case label="r2" cond="TQ1.selected.label in ['B', 'F']">의 시설, 장비, 제품 등은 목적을 달성하는 과정에 필요한 수단을 제공한다.</case>
  <case label="r3" cond="TQ1.selected.label in ['C', 'G']">의 서비스를 통해 목적을 달성할 수 있다.</case>
  <case label="r4" cond="TQ1.selected.label in ['D', 'H']">의 서비스는 목적을 달성하는 과정에 필요한 수단을 제공한다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<pipe
  label="Q1_row3_pipe"
  capture="">
  <case label="r1" cond="TQ1.selected.label in ['A', 'B', 'E', 'F']">에서 제공한 시설, 장비, 제품 등을 전 과정에 걸쳐 사용하는 데에 문제가 없다.</case>
  <case label="r2" cond="TQ1.selected.label in ['C', 'D', 'G', 'H']">에서 제공한 서비스는 전 과정에 걸쳐 이용하는 데에 문제가 없다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="Q1"
  surveyDisplay="desktop">
  <title>문 1】 먼저 <strong>${{res.pcsi_name}}</strong>에서 제공하는 서비스의 핵심내용 등에 대해 질문 드리겠습니다.<br/>[pipe: Q1_pipe]</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>전혀<br/>그렇지<br/>않다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>이다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>매우<br/>그렇다</col>
  <row label="_1" value="1"><strong>${{res.pcsi_name}}</strong>[pipe: Q1_row1_pipe]</row>
  <row label="_2" value="2"><strong>${{res.pcsi_name}}</strong>[pipe: Q1_row2_pipe]</row>
  <row label="_3" value="3"><strong>${{res.pcsi_name}}</strong>[pipe: Q1_row3_pipe]</row>
</radio>
<suspend/>

<pipe
  label="Q2_row3_pipe"
  capture="">
  <case label="r1" cond="TQ1.selected.label in ['A', 'B']">직원은 내가 필요한 시설, 장비, 제품 등을 이용할 수 있도록 최선을 다한다.</case>
  <case label="r2" cond="TQ1.selected.label in ['C', 'D']">직원은 내가 필요한 서비스를 받을 수 있도록 최선을 다한다.</case>
  <case label="r3" cond="TQ1.selected.label in ['E', 'F']">직원은 내 입장을 이해하고 시설, 장비, 제품 등을 이용할 수 있도록 최선을 다한다.</case>
  <case label="r4" cond="TQ1.selected.label in ['G', 'H']">직원은 내 입장을 이해하고 서비스를 받을 수 있도록 최선을 다한다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="Q2"
  surveyDisplay="desktop">
  <title>문 2】 다음은 <strong>${{res.pcsi_name}}</strong>의 서비스 제공 과정에 대한 질문입니다. 앞서 응답하신 방법대로 각 질문에 대해 고객님께서 동의하시는 정도에 따라 보기(11개) 중에서 골라주세요.</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>전혀<br/>그렇지<br/>않다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>이다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>매우<br/>그렇다</col>
  <row label="_1" value="1"><strong>${{res.pcsi_name}}</strong> 직원이 고객을 대하는 태도는 친절하고 만족스럽다.</row>
  <row label="_2" value="2"><strong>${{res.pcsi_name}}</strong>은 필요한 정보나 업무처리결과 등을 분명하게 알려준다.</row>
  <row label="_3" value="3"><strong>${{res.pcsi_name}}</strong> [pipe: Q2_row3_pipe]</row>
</radio>
<suspend/>

<radio
  label="Q3X1">
  <title>문 3-1】 선생님께서는 <strong>${{res.pcsi_name}}</strong>에 방문한 경험이 있으신가요?</title>
  <comment></comment>
  <row label="r1" value="1">예</row>
  <row label="r2" value="2">아니오</row>
</radio>
<suspend/>

<pipe
  label="Q3_row1_pipe"
  capture="">
  <case label="r1" cond="Q3X1.r1">관련 시설 및 환경은 쾌적하다.</case>
  <case label="r2" cond="Q3X1.r2">서비스는 쉽고 편하게 이용할 수 있다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="Q3"
  surveyDisplay="desktop">
  <title>문 3】 다음은 <strong>${{res.pcsi_name}}</strong>의 서비스 제공 환경에 대한 질문입니다. 앞서 응답하신 방법대로 각 질문에 대해 고객님께서 동의하시는 정도에 따라 보기(11개) 중에서 골라주세요.</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>전혀<br/>그렇지<br/>않다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>이다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>매우<br/>그렇다</col>
  <row label="_1" value="1"><strong>${{res.pcsi_name}}</strong>의 [pipe: Q3_row1_pipe]</row>
  <row label="_2" value="2"><strong>${{res.pcsi_name}}</strong> 직원은 고객의 요구사항을 신속하게 처리한다.</row>
  <row label="_3" value="3"><strong>${{res.pcsi_name}}</strong>의 직원의 업무처리 능력은 우수하다.</row>
</radio>
<suspend/>


<radio
  label="Q4"
  surveyDisplay="desktop">
  <title>문 4】 다음은 <strong>${{res.pcsi_name}}</strong>이 공공기관으로서 수행하는 사회적 책임과 역할에 대한 질문입니다. 앞서 응답하신 방법대로 각 질문에 대해 고객님께서 동의하시는 정도에 따라 보기(11개) 중에서 골라주세요.</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>전혀<br/>그렇지<br/>않다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>이다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>매우<br/>그렇다</col>
  <row label="_1" value="1"><strong>${{res.pcsi_name}}</strong>은 본래의 설립목적과 취지에 맞추어 정책을 추진하고 있다.</row>
  <row label="_2" value="2"><strong>${{res.pcsi_name}}</strong>은 국가나 사회의 미래가치 창출을 위한 업무를 수행하고 있다.</row>
  <row label="_3" value="3"><strong>${{res.pcsi_name}}</strong>의 업무수행은 공정하고 투명하게 이루어진다.</row>
</radio>
<suspend/>




<block label="Q5_block">

<pipe
  label="Q5_pre_pipe"
  capture="">
  <case label="r1" cond="KQ.r1">0점(매우 나쁘다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 좋다)까지 긍정적일수록 높은 점수, 부정적 일수록 낮은 점수를 주시면 됩니다.</case>
  <case label="r2" cond="KQ.r2">각 질문에 대해 고객님께서 동의하시는 정도에 따라 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점 중에서 골라주세요. 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<html label="Q5_pre">※ 다음은 고객님께서 <strong>${{res.pcsi_name}}</strong>을 통해 경험하신 서비스의 비교평가에 대한 질문입니다. [pipe: Q5_pre_pipe]</html>

<radio
  label="Q5_1">
  <title>문 5-1】 고객님께서 <strong>'사전에 기대하셨던 것’</strong>과 비교할 때, <strong>${{res.pcsi_name}}</strong>의 서비스는 어떠셨습니까?</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>기대보다<br/>매우<br/>나쁘다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>똑같다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)기대보다<br/>매우<br/>좋다</col>
</radio>


<radio
  label="Q5_2">
  <title>문 5-2】 고객님께서 상상하실 수 있는 <strong>‘가장 이상적인 서비스 수준’</strong>과 비교할 때, <strong>${{res.pcsi_name}}</strong>의 서비스는 어떠셨습니까?</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>이상치에<br/>매우<br/>못 미친다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>(중간)</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>이상치에<br/>매우<br/>가깝다</col>
</radio>

<radio
  label="Q5_3">
  <title>문 5-3】 고객님께서 경험하셨던 <strong>‘다른 기관이나 기업의 서비스’</strong>와 비교할 때, <strong>${{res.pcsi_name}}</strong>의 서비스는 어떠셨습니까?</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>다른<br/>기관(기업)보다<br/>매우 나쁘다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>똑같다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>다른<br/>기관(기업)보다<br/>매우 좋다</col>
</radio>

</block>
<note>Q5_block END</note>
<suspend/>


<pipe
  label="Q6_pipe"
  capture="">
  <case label="r1" cond="KQ.r1">0점(전혀 그렇지 않다) 부터 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점(매우 그렇다)까지 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.</case>
  <case label="r2" cond="KQ.r2">각 질문에 대해 고객님께서 동의하시는 정도에 따라 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10점 중에서 골라주세요. 동의하시는 정도가 클수록 높은 점수, 작을수록 낮은 점수를 주시면 됩니다.</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="Q6"
  surveyDisplay="desktop">
  <title>※ 다음은 고객님께서 <strong>${{res.pcsi_name}}</strong>의 서비스 이용경험 후 느끼신 전반적인 평가에 대한 질문입니다.<br/>[pipe: Q6_pipe]</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>전혀<br/>그렇지<br/>않다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>이다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>매우<br/>그렇다</col>
  <row label="_1" value="1"><strong>${{res.pcsi_name}}</strong>의 서비스가 전반적으로 만족스럽다.</row>
  <row label="_2" value="2"><strong>${{res.pcsi_name}}</strong>의 서비스를 이용한 후, 기관에 대해 긍정적인 느낌이 든다.</row>
</radio>
<suspend/>


<pipe
  label="Q7_pipe"
  capture="">
  <case label="r1" cond="QQQ12.r1">의 사업/활동으로 <strong>국민의 삶의 질</strong>이 향상되었다고 생각하십니까?</case>
  <case label="r2" cond="QQQ12.r2">의 서비스를 이용하신 후, 기관의 사업/활동에 대해 <strong>신뢰감을 갖게</strong>되었습니까?</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio
  label="Q7">
  <title>문 7】 고객님께서는 <strong>${{res.pcsi_name}}</strong>[pipe: Q7_pipe]</title>
  <comment></comment>
  <col label="c0" value="0">(0)<br/>전혀<br/>그렇지<br/>않다</col>
  <col label="c1" value="1">(1)</col>
  <col label="c2" value="2">(2)</col>
  <col label="c3" value="3">(3)</col>
  <col label="c4" value="4">(4)</col>
  <col label="c5" value="5">(5)<br/>보통<br/>이다</col>
  <col label="c6" value="6">(6)</col>
  <col label="c7" value="7">(7)</col>
  <col label="c8" value="8">(8)</col>
  <col label="c9" value="9">(9)</col>
  <col label="c10" value="10">(10)<br/>매우<br/>그렇다</col>
</radio>
<suspend/>

{xml_Q8_after[division]}


 <radio 
   label="DQ1"
   randomize="0">
    <title><strong>통계처리를 위해 추가적으로 몇 가지만 더 여쭈어 보겠습니다.</strong><br/><br/>
DQ1. 고객님의 성별은 어떻게 되십니까? </title>
    <row label="r1" value="1">남자</row>
    <row label="r2" value="2">여자</row>
  </radio>

  <suspend/>

{xml_DQ2[division]}

  <radio 
   label="DQ3"
   randomize="0">
    <title>DQ3. 고객님께서는 공공기관 고객만족도 조사와 관련하여 최근 3개월 이내에 예술의전당으로부터 거절할 수 없을 정도로 유리한 응답을 해 달라는 요청을 받으신 경험이 있으십니까? </title>
    <row label="r1" value="1">있다</row>
    <row label="r2" value="2">없다</row>
  </radio>

  <suspend/>

<pipe
  label="BB_pipe"
  capture="">
  <case label="r1" cond="QQQ12.r1">응답자 소재지</case>
  <case label="r2" cond="QQQ12.r2">법인 소재지</case>
  <case label="null" cond="1">UNDEFINED</case>
</pipe>

<radio 
  label="BB"
  randomize="0">
  <title><strong>[pipe: BB_pipe]</strong></title>
  <row label="r1" value="1">서울</row>
  <row label="r2" value="2">부산</row>
  <row label="r3" value="3">대구</row>
  <row label="r4" value="4">인천</row>
  <row label="r5" value="5">광주</row>
  <row label="r6" value="6">대전</row>
  <row label="r7" value="7">울산</row>
  <row label="r8" value="8">경기</row>
  <row label="r9" value="9">강원</row>
  <row label="r10" value="10">충남</row>
  <row label="r11" value="11">충북</row>
  <row label="r12" value="12">세종</row>
  <row label="r13" value="13">전남</row>
  <row label="r14" value="14">전북</row>
  <row label="r15" value="15">경남</row>
  <row label="r16" value="16">경북</row>
  <row label="r17" value="17">제주</row>
</radio>

<suspend/>

<text
  label="RespData"
  size="40"
  optional="0">
  <title><strong>마지막으로 고객님의 성함과 연락처는 어떻게 되십니까? <br />본 정보는 추후 검증을 위해 활용할 뿐 외부에 공개되지 않으며 검증 이후 일괄 폐기할 예정입니다.</strong></title>
  <comment></comment>
    <validate>
if len(this.phone.val) le 10 :
 error('최소 9자 이상으로 입력해주십시오.')
    </validate>
<style name='el.text' rows="name"> <![CDATA[
<div style="display: flex; flex-direction: row;">
  <div>
    <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 20 20" fill="currentColor" style="width:25px; margin:0 auto;">
      <path fill-rule="evenodd" d="M18 10a8 8 0 11-16 0 8 8 0 0116 0zm-6-3a2 2 0 11-4 0 2 2 0 014 0zm-2 4a5 5 0 00-4.546 2.916A5.986 5.986 0 0010 16a5.986 5.986 0 004.546-2.084A5 5 0 0010 11z" clip-rule="evenodd" />
    </svg>
  </div>
  <div>
    <input type="text" name="$(name)" id="$(id)" value="$(value)" size="$(size)" class="input text-input" $(extra) placeholder="${{row.styles.ss.preText or this.styles.ss.preText or ''}}"/>
  </div>
</div>
]]></style>
<style name='el.text' rows="phone"> <![CDATA[
<div style="display: flex; flex-direction: row;">
  <div>
    <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 20 20" fill="currentColor" style="width:25px; margin:0 auto;">
      <path fill-rule="evenodd" d="M7 2a2 2 0 00-2 2v12a2 2 0 002 2h6a2 2 0 002-2V4a2 2 0 00-2-2H7zm3 14a1 1 0 100-2 1 1 0 000 2z" clip-rule="evenodd" />
    </svg>
  </div>
  <div>
    <input type="text" name="$(name)" id="$(id)" value="$(value)" size="$(size)" class="input text-input" $(extra) placeholder="${{row.styles.ss.preText or this.styles.ss.preText or ''}}"/>
  </div>
</div>
]]></style>
    <style name="question.after" wrap="ready"><![CDATA[
phoneInput();

function phoneInput(){{
$ ('.phone').keyup(function(){{
$ ('.phone').find('input[type=text]').val( $ ('.phone').find('input[type=text]').val().replace(/[^0-9]/g, "").replace(/(^02|^0505|^1[0-9]{{3}}|^0[0-9]{{2}})([0-9]+)?([0-9]{{4}})$/,"$1-$2-$3").replace("--", "-") );
}}).trigger('keyup');
}}
]]></style>
  <row label="name" ss:preText="성명"/>
  <row label="phone" ss:rowClassNames="phone" ss:preText="연락처 (숫자만 입력해주십시오.)"/>
</text>



<suspend/>

<exec>
status(True,'r1')
</exec>
</survey>
    '''
    
    xml_filename = 'survey.xml'
    with open(os.path.join(os.getcwd(), save_folder, xml_filename), 'w', encoding='utf-8') as f :
        f.write(xml)
